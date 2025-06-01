from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QLabel, QTableWidget, QTableWidgetItem,
    QHeaderView, QHBoxLayout, QLineEdit, QComboBox, QPushButton, QGroupBox
)
from PyQt5.QtCore import Qt,QDate
from db_helper import get_student_term_summary_rows
import jdatetime
from shamsi_date_picker import ShamsiDatePicker
from shamsi_date_popup import ShamsiDatePopup
from datetime import datetime, timedelta

class StudentTermSummaryWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("گزارش کلی هنرجویان")
        self.setGeometry(300, 150, 1200, 600)

        layout = QVBoxLayout()
        layout.addWidget(self.create_filter_box())
        layout.addWidget(self.create_table())

        self.summary_label = QLabel("تعداد نتایج: ۰")
        layout.addWidget(self.summary_label)

        self.setLayout(layout)
        self.load_filter_options()

        self.set_default_dates()

        self.load_data(apply_filters=False)
        self.showMaximized()
        self.table.setSortingEnabled(True)


    def create_filter_box(self):
        group = QGroupBox("فیلترها")
        layout = QHBoxLayout()

        # فیلترهای متنی
        self.input_student = QLineEdit()
        self.input_student.setPlaceholderText("نام هنرجو")

        # ComboBoxها
        self.combo_teacher = QComboBox()
        self.combo_teacher.addItem("همه", None)

        self.combo_instrument = QComboBox()
        self.combo_instrument.addItem("همه", None)

        self.combo_class = QComboBox()
        self.combo_class.addItem("همه", None)

        self.combo_day = QComboBox()
        self.combo_day.addItem("همه", None)
        self.combo_day.addItems(["شنبه", "یکشنبه", "دوشنبه", "سه‌شنبه", "چهارشنبه", "پنج‌شنبه", "جمعه"])

        # تاریخ شروع و پایان
        self.date_from = ShamsiDatePicker()
        self.date_to = ShamsiDatePicker()

        # فیلتر فعال / پایان یافته
        self.combo_term_status = QComboBox()
        self.combo_term_status.addItems(["همه ترم‌ها", "فقط فعال", "فقط پایان یافته"])

        # دکمه فیلتر
        self.btn_filter = QPushButton("اعمال فیلتر")
        self.btn_filter.clicked.connect(lambda: self.load_data(apply_filters=True))


        # چیدن در layout
        layout.addWidget(QLabel("🎓 هنرجو:"))
        layout.addWidget(self.input_student)
        layout.addWidget(QLabel("👨‍🏫 استاد:"))
        layout.addWidget(self.combo_teacher)
        layout.addWidget(QLabel("🎼 ساز:"))
        layout.addWidget(self.combo_instrument)
        layout.addWidget(QLabel("🏫 کلاس:"))
        layout.addWidget(self.combo_class)
        layout.addWidget(QLabel("🗓️ روز:"))
        layout.addWidget(self.combo_day)
        layout.addWidget(QLabel("از تاریخ:"))
        layout.addWidget(self.date_from)
        layout.addWidget(QLabel("تا تاریخ:"))
        layout.addWidget(self.date_to)
        layout.addWidget(QLabel("ترم:"))
        layout.addWidget(self.combo_term_status)
        layout.addWidget(self.btn_filter)
        self.btn_clear = QPushButton("🧹 پاکسازی فیلترها")
        self.btn_clear.clicked.connect(self.clear_filters)
        layout.addWidget(self.btn_clear)
        self.btn_export = QPushButton("📤 خروجی اکسل")
        self.btn_export.clicked.connect(self.export_to_excel)
        layout.addWidget(self.btn_export)


        group.setLayout(layout)
        return group


    def create_table(self):
        self.table = QTableWidget()
        self.table.setColumnCount(13)
        self.table.setHorizontalHeaderLabels([
            "نام هنرجو", "کد ملی", "نام کلاس", "نام استاد", "ساز", "روز",
            "ساعت شروع", "تاریخ شروع ترم", "تاریخ پایان ترم",
            "تعداد جلسات", "حضور", "غیبت", "نسبت حضور (%)"
        ])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        return self.table



    def load_data(self,apply_filters=False):

        if not apply_filters:
            # فقط همه داده‌ها رو بدون هیچ فیلتر نمایش بده
            rows = get_student_term_summary_rows()
        else:
            student_name = self.input_student.text().strip()

            teacher_name = self.combo_teacher.currentText()
            if teacher_name == "همه": teacher_name = ""

            class_name = self.combo_class.currentText()
            if class_name == "همه": class_name = ""

            instrument_name = self.combo_instrument.currentText()
            if instrument_name == "همه": instrument_name = ""

            day = self.combo_day.currentText()
            if day == "همه": day = ""

            date_from = self.date_from.selected_shamsi.strip()
            date_to = self.date_to.selected_shamsi.strip()

            term_status = self.combo_term_status.currentText()
            if term_status == "فقط فعال":
                term_status = "active"
            elif term_status == "فقط پایان یافته":
                term_status = "finished"
            else:
                term_status = ""

            rows = get_student_term_summary_rows(
                student_name=student_name,
                teacher_name=teacher_name,
                class_name=class_name,
                instrument_name=instrument_name,
                day=day,
                date_from=date_from,
                date_to=date_to,
                term_status=term_status
            )

        self.table.setRowCount(len(rows))
        for row_idx, row_data in enumerate(rows):
            for col_idx, value in enumerate(row_data):
                item = QTableWidgetItem(str(value))
                item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(row_idx, col_idx, item)

        self.summary_label.setText(f"تعداد نتایج: {len(rows)}")

    def load_filter_options(self):
        from db_helper import fetch_teachers_simple, fetch_classes

        self.combo_teacher.addItems([t[1] for t in fetch_teachers_simple()])
        self.combo_class.addItems([c[1] for c in fetch_classes()])

        # سازها را از کلاس‌ها استخراج می‌کنیم (تکراری‌ها حذف می‌شوند)
        instruments = set()
        for c in fetch_classes():
            instruments.add(c[3])  # فیلد instrument
        self.combo_instrument.addItems(sorted(instruments))

    def clear_filters(self):
        self.input_student.clear()
        self.combo_teacher.setCurrentIndex(0)
        self.combo_class.setCurrentIndex(0)
        self.combo_instrument.setCurrentIndex(0)
        self.combo_day.setCurrentIndex(0)
        self.combo_term_status.setCurrentIndex(0)

        self.set_default_dates()

        # نمایش مجدد همه داده‌ها بدون فیلتر
        self.load_data(apply_filters=False)

    def set_default_dates(self):
        today = QDate.currentDate()
        three_months_ago = today.addMonths(-3)

        self.date_from.setDate(three_months_ago)
        self.date_to.setDate(today)

    def export_to_excel(self):
        from PyQt5.QtWidgets import QFileDialog
        import openpyxl
        import jdatetime

        today = jdatetime.date.today().strftime("%Y%m%d")
        default_name = f"گزارش_کلی_هنرجویان_{today}.xlsx"

        file_path, _ = QFileDialog.getSaveFileName(self, "ذخیره فایل اکسل", default_name, "Excel Files (*.xlsx)")
        if not file_path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Student Term Summary"

        # هدر جدول
        for col in range(self.table.columnCount()):
            header = self.table.horizontalHeaderItem(col)
            if header:
                ws.cell(row=1, column=col + 1, value=header.text())

        # داده‌ها
        for row in range(self.table.rowCount()):
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if item:
                    ws.cell(row=row + 2, column=col + 1, value=item.text())

        wb.save(file_path)
