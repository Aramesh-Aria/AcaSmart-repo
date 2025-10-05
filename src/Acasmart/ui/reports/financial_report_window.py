from acasmart.data.repos.reports_repo import get_all_student_terms_with_financials
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QLabel, QTableWidget, QTableWidgetItem, QHeaderView,
    QHBoxLayout, QLineEdit, QComboBox, QPushButton, QFileDialog
)
from acasmart.data.repos.classes_repo import fetch_classes
from PySide6.QtCore import Qt, QDate
from PySide6.QtGui import QColor

from acasmart.core.utils import format_currency_with_unit
from acasmart.ui.widgets.shamsi_date_picker import ShamsiDatePicker

import jdatetime
import openpyxl


class FinancialReportWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("گزارش مالی هنرجویان")
        self.setGeometry(300, 200, 1300, 500)
        self.all_data = []  # همه داده‌ها ذخیره می‌شن برای فیلتر کردن
        self.build_ui()

    def build_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(10)

        title = QLabel("📊 گزارش مالی ترم‌های هنرجویان")
        title.setStyleSheet("font-size: 16px; font-weight: bold;")
        layout.addWidget(title)

        # ویجت خلاصه آماری
        self.summary_label = QLabel("")
        self.summary_label.setStyleSheet("font-size: 14px; font-weight: bold; color: #333;")
        layout.addWidget(self.summary_label)

        # فیلتر بالا
        filter_layout = QHBoxLayout()
        self.input_student_name = QLineEdit()
        self.input_student_name.setPlaceholderText("نام هنرجو")

        self.combo_class = QComboBox()
        self.combo_class.addItem("همه کلاس‌ها", None)
        for cid, cname, *_ in fetch_classes():
            self.combo_class.addItem(cname, cid)

        self.combo_status = QComboBox()
        self.combo_status.addItems(["همه وضعیت‌ها", "تسویه", "بدهکار"])

        btn_filter = QPushButton("اعمال فیلتر")
        btn_filter.clicked.connect(self.apply_filters)

        btn_clear = QPushButton("پاکسازی فیلتر")
        btn_clear.clicked.connect(self.reset_filters)

        btn_export = QPushButton("📤 خروجی اکسل")
        btn_export.clicked.connect(self.export_to_excel)

        filter_layout.addWidget(self.input_student_name)
        filter_layout.addWidget(self.combo_class)
        filter_layout.addWidget(self.combo_status)
        filter_layout.addWidget(btn_filter)
        filter_layout.addWidget(btn_clear)
        filter_layout.addWidget(btn_export)

        layout.addLayout(filter_layout)

        # فیلتر بازه تاریخ ترم
        self.date_from_picker = ShamsiDatePicker(": از تاریخ")
        self.date_to_picker = ShamsiDatePicker(": تا تاریخ")

        today = QDate.currentDate()
        three_months_ago = today.addMonths(-3)
        self.date_from_picker.setDate(three_months_ago)
        self.date_to_picker.setDate(today)

        filter_layout.addWidget(self.date_from_picker)
        filter_layout.addWidget(self.date_to_picker)

        # جدول گزارش
        self.table = QTableWidget()
        self.table.setColumnCount(10)
        self.table.setHorizontalHeaderLabels([
            "نام هنرجو", "کلاس", "ساز", "استاد", "شروع ترم", "پایان ترم",
            "شهریه", "پرداخت شهریه", "بدهی", "وضعیت مالی"
        ])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.table)
        self.table.setSortingEnabled(True)
        self.showMaximized()

        self.setLayout(layout)
        self.load_data()

    def load_data(self):
        self.all_data = get_all_student_terms_with_financials()
        self.populate_table(self.all_data)

    def populate_table(self, data):
        self.table.setRowCount(len(data))
        total_tuition = 0
        total_paid = 0
        total_debt = 0

        for i, row in enumerate(data):
            self.table.setItem(i, 0, QTableWidgetItem(row['student_name']))
            self.table.setItem(i, 1, QTableWidgetItem(row['class_name']))
            self.table.setItem(i, 2, QTableWidgetItem(row['instrument']))
            self.table.setItem(i, 3, QTableWidgetItem(row['teacher_name']))
            self.table.setItem(i, 4, QTableWidgetItem(row['start_date']))
            self.table.setItem(i, 5, QTableWidgetItem(row['end_date'] or "—"))
            self.table.setItem(i, 6, QTableWidgetItem(format_currency_with_unit(row['tuition'])))
            self.table.setItem(i, 7, QTableWidgetItem(format_currency_with_unit(row['paid_tuition'])))
            self.table.setItem(i, 8, QTableWidgetItem(format_currency_with_unit(row['debt'])))

            item_status = QTableWidgetItem(row['status'])

            if row['status'] == "تسویه":
                item_status.setBackground(QColor("#81C784"))  # سبز مثل نمونه «شهریه»
                item_status.setForeground(QColor("#0B3D17"))
            elif row['status'] == "بدهکار":
                item_status.setBackground(QColor("#E57373"))  # قرمز ملایم
                item_status.setForeground(QColor("#5D1919"))

            item_status.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(i, 9, item_status)

            total_tuition += row['tuition']
            total_paid += row['paid_tuition']
            total_debt += row['debt']

        self.summary_label.setText(
            f"تعداد ترم‌ها: {len(data)}   |   مجموع شهریه: {format_currency_with_unit(total_tuition)}   |   مجموع پرداخت: {format_currency_with_unit(total_paid)}   |   مجموع بدهی: {format_currency_with_unit(total_debt)}"
        )
        self.filtered_data = data

    def apply_filters(self):
        name_filter = self.input_student_name.text().strip()
        class_id = self.combo_class.currentData()
        status = self.combo_status.currentText()
        from_date = self.date_from_picker.get_miladi_str()
        to_date = self.date_to_picker.get_miladi_str()

        filtered = []
        for row in self.all_data:
            if name_filter and name_filter not in row['student_name']:
                continue
            if class_id and row['class_id'] != class_id:
                continue
            if status != "همه وضعیت‌ها" and row['status'] != status:
                continue

            term_start = row['start_date']
            term_end = row['end_date'] or "2100-01-01"
            if not (term_start <= to_date and term_end >= from_date):
                continue

            filtered.append(row)

        self.populate_table(filtered)

    def reset_filters(self):
        self.input_student_name.clear()
        self.combo_class.setCurrentIndex(0)
        self.combo_status.setCurrentIndex(0)

        self.date_from_picker.setDate(QDate.currentDate().addMonths(-3))
        self.date_to_picker.setDate(QDate.currentDate())

        self.populate_table(self.all_data)

    def export_to_excel(self):
        today_shamsi = jdatetime.date.today().strftime("%Y-%m-%d")
        suggested_filename = f"گزارش مالی - {today_shamsi}.xlsx"

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "ذخیره فایل اکسل",
            suggested_filename,
            "Excel Files (*.xlsx)"
        )
        if not file_path:
            return

        headers = [
            "نام هنرجو", "کلاس", "ساز", "استاد", "شروع ترم", "پایان ترم",
            "شهریه", "پرداخت شهریه", "بدهی", "وضعیت مالی"
        ]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Financial Report"

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        for row_idx, row_data in enumerate(self.filtered_data, start=2):
            ws.cell(row=row_idx, column=1, value=row_data['student_name'])
            ws.cell(row=row_idx, column=2, value=row_data['class_name'])
            ws.cell(row=row_idx, column=3, value=row_data['instrument'])
            ws.cell(row=row_idx, column=4, value=row_data['teacher_name'])
            ws.cell(row=row_idx, column=5, value=row_data['start_date'])
            ws.cell(row=row_idx, column=6, value=row_data['end_date'] or "—")
            ws.cell(row=row_idx, column=7, value=row_data['tuition'])
            ws.cell(row=row_idx, column=8, value=row_data['paid_tuition'])
            ws.cell(row=row_idx, column=9, value=row_data['debt'])
            ws.cell(row=row_idx, column=10, value=row_data['status'])

        wb.save(file_path)