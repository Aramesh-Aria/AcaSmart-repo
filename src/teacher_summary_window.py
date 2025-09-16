from data.reports_repo import get_teacher_summary_rows
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QTableWidget, QTableWidgetItem,
    QHeaderView, QPushButton, QFileDialog, QComboBox, QSizePolicy
)
from PySide6.QtCore import Qt
from data.teachers_repo import fetch_teachers_simple
import openpyxl
import jdatetime


class TeacherSummaryWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Ú¯Ø²Ø§Ø±Ø´ Ú©Ù„ÛŒ Ø§Ø³Ø§ØªÛŒØ¯")
        self.setGeometry(300, 150, 1100, 600)

        layout = QVBoxLayout()
        layout.addLayout(self.create_filter_box())

        self.table = QTableWidget()
        self.table.setColumnCount(9)
        self.table.setHorizontalHeaderLabels([
            "Ù†Ø§Ù… Ø§Ø³ØªØ§Ø¯", "Ú©Ø¯ Ù…Ù„ÛŒ", "Ú©Ø§Ø±Øª ØªØ¯Ø±ÛŒØ³", "ØªÙ„ÙÙ†", "ØªØ§Ø±ÛŒØ® ØªÙˆÙ„Ø¯",
            "Ø´Ù…Ø§Ø±Ù‡ Ú©Ø§Ø±Øª", "Ø´Ù…Ø§Ø±Ù‡ Ø´Ø¨Ø§", "Ø³Ø§Ø²Ù‡Ø§ÛŒ ØªØ¯Ø±ÛŒØ³ÛŒ", "Ø±ÙˆØ²Ù‡Ø§ÛŒ Ú©Ù„Ø§Ø³"
        ])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.table)

        self.summary_label = QLabel()
        layout.addWidget(self.summary_label)

        self.setLayout(layout)
        self.load_filter_options()
        self.load_data()
        self.showMaximized()
        self.table.setSortingEnabled(True)

    def create_filter_box(self):
        layout = QHBoxLayout()

        self.combo_teacher = QComboBox()
        self.combo_teacher.addItem("Ù‡Ù…Ù‡", None)

        self.combo_day = QComboBox()
        self.combo_day.addItem("Ù‡Ù…Ù‡", None)
        self.combo_day.addItems(["Ø´Ù†Ø¨Ù‡", "ÛŒÚ©Ø´Ù†Ø¨Ù‡", "Ø¯ÙˆØ´Ù†Ø¨Ù‡", "Ø³Ù‡â€ŒØ´Ù†Ø¨Ù‡", "Ú†Ù‡Ø§Ø±Ø´Ù†Ø¨Ù‡", "Ù¾Ù†Ø¬â€ŒØ´Ù†Ø¨Ù‡", "Ø¬Ù…Ø¹Ù‡"])

        btn_filter = QPushButton("Ø§Ø¹Ù…Ø§Ù„ ÙÛŒÙ„ØªØ±")
        btn_filter.clicked.connect(lambda: self.load_data(apply_filters=True))

        btn_clear = QPushButton("ğŸ§¹ Ù¾Ø§Ú©Ø³Ø§Ø²ÛŒ ÙÛŒÙ„ØªØ±Ù‡Ø§")
        btn_clear.clicked.connect(self.clear_filters)

        export_btn = QPushButton("ğŸ“¤ Ø®Ø±ÙˆØ¬ÛŒ Ø§Ú©Ø³Ù„")
        export_btn.clicked.connect(self.export_to_excel)

        for widget in [self.combo_teacher, self.combo_day, btn_filter, btn_clear, export_btn]:
            widget.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)

        layout.addWidget(QLabel("ğŸ‘¨â€ğŸ« Ø§Ø³ØªØ§Ø¯:"))
        layout.addWidget(self.combo_teacher)
        layout.addSpacing(10)
        layout.addWidget(QLabel("ğŸ—“ï¸ Ø±ÙˆØ² ØªØ¯Ø±ÛŒØ³:"))
        layout.addWidget(self.combo_day)
        layout.addSpacing(10)
        layout.addWidget(btn_filter)
        layout.addSpacing(10)
        layout.addWidget(btn_clear)
        layout.addStretch()
        layout.addWidget(export_btn)

        return layout

    def load_filter_options(self):
        self.combo_teacher.addItems([t[1] for t in fetch_teachers_simple()])

    def load_data(self, apply_filters=False):
        rows = get_teacher_summary_rows()

        if apply_filters:
            teacher_name = self.combo_teacher.currentText()
            if teacher_name == "Ù‡Ù…Ù‡":
                teacher_name = ""

            selected_day = self.combo_day.currentText()
            if selected_day == "Ù‡Ù…Ù‡":
                selected_day = ""

            filtered = []
            for row in rows:
                match_teacher = teacher_name in row[0] if teacher_name else True
                match_day = selected_day in row[8] if selected_day else True

                if match_teacher and match_day:
                    filtered.append(row)
            rows = filtered

        self.table.setRowCount(len(rows))
        for row_idx, row_data in enumerate(rows):
            for col_idx, value in enumerate(row_data):
                item = QTableWidgetItem(str(value) if value else "â€”")
                item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(row_idx, col_idx, item)

        self.summary_label.setText(f"ØªØ¹Ø¯Ø§Ø¯ Ø§Ø³Ø§ØªÛŒØ¯: {len(rows)}")

    def clear_filters(self):
        self.combo_teacher.setCurrentIndex(0)
        self.combo_day.setCurrentIndex(0)
        self.load_data(apply_filters=False)

    def export_to_excel(self):
        today = jdatetime.date.today().strftime("%Y%m%d")
        default_name = f"Ú¯Ø²Ø§Ø±Ø´_Ø§Ø³Ø§ØªÛŒØ¯_{today}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(self, "Ø°Ø®ÛŒØ±Ù‡ ÙØ§ÛŒÙ„ Ø§Ú©Ø³Ù„", default_name, "Excel Files (*.xlsx)")

        if not file_path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Teachers Summary"

        for col in range(self.table.columnCount()):
            header = self.table.horizontalHeaderItem(col)
            if header:
                ws.cell(row=1, column=col + 1, value=header.text())

        for row in range(self.table.rowCount()):
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if item:
                    ws.cell(row=row + 2, column=col + 1, value=item.text())

        wb.save(file_path)