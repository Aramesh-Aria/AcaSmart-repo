from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QLabel, QTableWidget, QTableWidgetItem, QHeaderView,
    QLineEdit, QComboBox, QPushButton, QHBoxLayout, QFileDialog, QMessageBox
)
from PySide6.QtCore import Qt, QDate

from db_helper import get_attendance_report_rows,fetch_classes
from datetime import datetime
import openpyxl
import jdatetime

from shamsi_date_picker import ShamsiDatePicker
from shamsi_date_popup import ShamsiDatePopup


class AttendanceReportWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("گزارش حضور و غیاب هنرجویان")
        self.setGeometry(250, 150, 1300, 600)
        self.build_ui()

    def build_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(10)

        title = QLabel("📅 گزارش کلی حضور و غیاب هنرجویان")
        title.setStyleSheet("font-size: 16px; font-weight: bold;")
        layout.addWidget(title)

        filter_layout = QHBoxLayout()

        self.input_student_name = QLineEdit()
        self.input_student_name.setPlaceholderText("نام هنرجو")

        self.combo_teacher = QComboBox()
        self.combo_teacher.addItem("همه اساتید", None)

        self.combo_class = QComboBox()
        self.combo_class.addItem("همه کلاس‌ها", None)

        self.combo_term_status = QComboBox()
        self.combo_term_status.addItem("همه ترم‌ها", None)
        self.combo_term_status.addItem("فقط فعال", "active")
        self.combo_term_status.addItem("فقط پایان‌یافته", "finished")

        btn_filter = QPushButton("اعمال فیلتر")
        btn_filter.clicked.connect(self.apply_filters)

        btn_clear = QPushButton("پاکسازی فیلتر")
        btn_clear.clicked.connect(self.reset_filters)

        btn_export = QPushButton("📤 خروجی اکسل")
        btn_export.clicked.connect(self.export_to_excel)

        filter_layout.addWidget(self.input_student_name)
        filter_layout.addWidget(self.combo_teacher)
        filter_layout.addWidget(self.combo_class)
        filter_layout.addWidget(self.combo_term_status)
        filter_layout.addWidget(btn_filter)
        filter_layout.addWidget(btn_clear)
        filter_layout.addWidget(btn_export)

        self.date_from_picker = ShamsiDatePicker(": از تاریخ ترم")
        self.date_to_picker = ShamsiDatePicker(": تا تاریخ ترم")

        today = QDate.currentDate()
        three_months_ago = today.addMonths(-3)
        self.date_from_picker.setDate(three_months_ago)
        self.date_to_picker.setDate(today)

        filter_layout.addWidget(self.date_from_picker)
        filter_layout.addWidget(self.date_to_picker)

        layout.addLayout(filter_layout)

        self.table = QTableWidget()
        layout.addWidget(self.table)

        self.status_label = QLabel("تعداد نتایج: ۰")
        layout.addWidget(self.status_label)

        self.showMaximized()
        self.setLayout(layout)
        self.load_data()
        self.table.setSortingEnabled(True)

    def load_data(self):
        rows = get_attendance_report_rows()
        self.all_data = rows
        self.combo_teacher.clear()
        self.combo_teacher.addItem("همه اساتید", None)

        self.combo_class.clear()
        self.combo_class.addItem("همه کلاس‌ها", None)
        for cid, cname, *_ in fetch_classes():
            self.combo_class.addItem(cname, cid)

        teacher_names = set()
        class_names = set()
        for row in rows:
            teacher_names.add(row['teacher_name'])
            class_names.add(row['class_name'])

        for t in sorted(teacher_names):
            self.combo_teacher.addItem(t, t)



        self.populate_table(self.all_data)
        self.status_label.setText(f"تعداد نتایج: {len(self.all_data)}")

    def apply_filters(self):
        name_filter = self.input_student_name.text().strip()
        teacher_filter = self.combo_teacher.currentData()
        class_filter = self.combo_class.currentData()
        term_status_filter = self.combo_term_status.currentData()
        from_date = self.date_from_picker.get_miladi_str()
        to_date = self.date_to_picker.get_miladi_str()

        filtered = []

        for row in self.all_data:
            if name_filter and name_filter not in row['student_name']:
                continue
            if teacher_filter and row['teacher_name'] != teacher_filter:
                continue
            if class_filter and row['class_id'] != class_filter:
                continue
            if term_status_filter == "active" and row['end_date']:
                continue
            if term_status_filter == "finished" and not row['end_date']:
                continue
            if row['start_date'] > to_date or (row['end_date'] and row['end_date'] < from_date):
                continue

            filtered.append(row)

        if not filtered:
            QMessageBox.information(self, "بدون نتیجه", "هیچ رکوردی مطابق با فیلترهای انتخاب‌شده یافت نشد.")

        self.populate_table(filtered)
        self.status_label.setText(f"تعداد نتایج: {len(filtered)}")

    def populate_table(self, data):
        all_dates = set()
        for row in data:
            start = row['start_date']
            end = row['end_date'] or "2100-01-01"
            for date in row['attendance'].keys():
                if start <= date <= end:
                    all_dates.add(date)
        sorted_dates = sorted(all_dates)

        headers = ["هنرجو", "استاد", "کلاس", "ساز", "شروع ترم", "پایان ترم"] + sorted_dates

        self.table.clearContents()
        self.table.setRowCount(0)
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)

        for col in range(6, len(headers)):
            item = self.table.horizontalHeaderItem(col)
            if not item:
                item = QTableWidgetItem(headers[col])
                self.table.setHorizontalHeaderItem(col, item)
            item.setBackground(Qt.lightGray)

        self.table.setRowCount(len(data))
        header = self.table.horizontalHeader()
        for col in range(len(headers)):
            if col == 0:
                header.setSectionResizeMode(col, QHeaderView.Stretch)
            elif 1 <= col <= 5:
                header.setSectionResizeMode(col, QHeaderView.ResizeToContents)
            else:
                self.table.setColumnWidth(col, 85)

        for i, row in enumerate(data):
            self.table.setItem(i, 0, QTableWidgetItem(row['student_name']))
            self.table.setItem(i, 1, QTableWidgetItem(row['teacher_name']))
            self.table.setItem(i, 2, QTableWidgetItem(row['class_name']))
            self.table.setItem(i, 3, QTableWidgetItem(row['instrument']))
            self.table.setItem(i, 4, QTableWidgetItem(row['start_date']))
            self.table.setItem(i, 5, QTableWidgetItem(row['end_date'] or "—"))

            for j, date in enumerate(sorted_dates, start=6):
                status = row['attendance'].get(date, "")
                item = QTableWidgetItem(status)
                if status == "حاضر":
                    item.setForeground(Qt.green)
                elif status == "غایب":
                    item.setForeground(Qt.red)
                self.table.setItem(i, j, item)

    def reset_filters(self):
        self.input_student_name.clear()
        self.combo_teacher.setCurrentIndex(0)
        self.combo_class.setCurrentIndex(0)
        self.combo_term_status.setCurrentIndex(0)

        self.date_from_picker.setDate(QDate.currentDate().addMonths(-3))
        self.date_to_picker.setDate(QDate.currentDate())

        self.populate_table(self.all_data)
        self.status_label.setText(f"تعداد نتایج: {len(self.all_data)}")

    def export_to_excel(self):
        today = jdatetime.date.today().strftime("%Y%m%d")
        default_name = f"گزارش_حضور_هنرجویان_{today}.xlsx"

        file_path, _ = QFileDialog.getSaveFileName(self, "ذخیره فایل اکسل", default_name, "Excel Files (*.xlsx)")
        if not file_path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        for col, header in enumerate(
                [self.table.horizontalHeaderItem(i).text() for i in range(self.table.columnCount())], start=1):
            ws.cell(row=1, column=col, value=header)

        for row in range(self.table.rowCount()):
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if item:
                    ws.cell(row=row + 2, column=col + 1, value=item.text())

        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

        wb.save(file_path)
